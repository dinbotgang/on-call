import csv
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Dallas Soccer Coaches"

# Colors
HEADER_BG     = "1A3A5C"   # dark navy
HEADER_FG     = "FFFFFF"
HS_BG         = "D6E4F0"   # light blue  – high school rows
CLUB_BG       = "EAF4EA"   # light green – club rows
ALT_HS_BG     = "C2D9EE"
ALT_CLUB_BG   = "D4EDD4"
ACCENT        = "2E86C1"   # medium blue for org column

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Read CSV
with open("/home/node/.openclaw/workspace/dallas_soccer_coaches.csv", newline="") as f:
    rows = list(csv.reader(f))

headers = rows[0]
data    = rows[1:]

# ── Column widths ──────────────────────────────────────────
col_widths = [14, 32, 24, 34, 36, 16, 40]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 28

# ── Header row ─────────────────────────────────────────────
header_font  = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
header_fill  = PatternFill("solid", fgColor=HEADER_BG)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col, val in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=val.upper())
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align
    cell.border    = border

# ── Data rows ──────────────────────────────────────────────
hs_count   = 0
club_count = 0

for row_idx, row in enumerate(data, 2):
    row_type = row[0] if row else ""
    is_hs    = row_type == "HIGH SCHOOL"

    if is_hs:
        hs_count += 1
        bg = HS_BG if hs_count % 2 == 1 else ALT_HS_BG
    else:
        club_count += 1
        bg = CLUB_BG if club_count % 2 == 1 else ALT_CLUB_BG

    fill = PatternFill("solid", fgColor=bg)

    ws.row_dimensions[row_idx].height = 18

    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill      = fill
        cell.border    = border
        cell.font      = Font(name="Calibri", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 7))

        # Highlight email column in blue
        if col_idx == 5 and val and "@" in val:
            cell.font = Font(name="Calibri", size=10, color="1565C0", underline="single")

        # Bold org name
        if col_idx == 2:
            cell.font = Font(name="Calibri", size=10, bold=True)

        # Type badge – color the TYPE cell text
        if col_idx == 1:
            color = "1A3A5C" if is_hs else "1B5E20"
            cell.font = Font(name="Calibri", size=10, bold=True, color=color)

# ── Freeze header + auto-filter ────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"

# ── Legend sheet ───────────────────────────────────────────
leg = wb.create_sheet("Legend")
leg.column_dimensions["A"].width = 20
leg.column_dimensions["B"].width = 40

legend_data = [
    ("Legend", ""),
    ("🔵 HIGH SCHOOL", "Public/private high school soccer coaches"),
    ("🟢 CLUB",        "Youth/competitive soccer club coaches"),
    ("",               ""),
    ("Tips", ""),
    ("Filter by Type", "Use dropdown arrow on column A"),
    ("Filter by Org",  "Use dropdown arrow on column B"),
    ("Emails",         "Shown in blue – click to copy"),
]

for r, (a, b) in enumerate(legend_data, 1):
    ca = leg.cell(row=r, column=1, value=a)
    cb = leg.cell(row=r, column=2, value=b)
    if a in ("Legend", "Tips"):
        ca.font = Font(bold=True, size=12, color=HEADER_BG)
    else:
        ca.font = Font(size=10)
        cb.font = Font(size=10)

wb.save("/home/node/.openclaw/workspace/dallas_soccer_coaches.xlsx")
print("Done!")
