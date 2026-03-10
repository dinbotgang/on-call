import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "HS & MS Soccer Coaches"

# Color palette
HEADER_BG   = "1A3A5C"
HEADER_FG   = "FFFFFF"
HS_BG       = "D6E4F0"
HS_ALT      = "BDD7EE"
MS_BG       = "E2EFDA"
MS_ALT      = "C6E0B4"

thin   = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["Level","ISD / District","School","Name","Role","Email","Phone","Source URL"]
col_widths = [14, 22, 32, 24, 32, 38, 16, 45]

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 28

# Header
for col, val in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=val.upper())
    c.font      = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
    c.fill      = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = border

with open("/home/node/.openclaw/workspace/hs_ms_soccer_coaches.csv", newline="") as f:
    rows = list(csv.reader(f))[1:]  # skip header

hs_count = ms_count = 0

for ri, row in enumerate(rows, 2):
    level = row[0] if row else ""
    is_hs = "HIGH SCHOOL" in level

    if is_hs:
        hs_count += 1
        bg = HS_BG if hs_count % 2 == 1 else HS_ALT
    else:
        ms_count += 1
        bg = MS_BG if ms_count % 2 == 1 else MS_ALT

    fill = PatternFill("solid", fgColor=bg)
    ws.row_dimensions[ri].height = 18

    for ci, val in enumerate(row, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.fill   = fill
        c.border = border
        c.font   = Font(name="Calibri", size=10)
        c.alignment = Alignment(vertical="center", wrap_text=(ci in (5, 8)))

        if ci == 1:  # Level badge
            color = "1A3A5C" if is_hs else "375623"
            c.font = Font(name="Calibri", size=10, bold=True, color=color)
        elif ci == 3:  # School name bold
            c.font = Font(name="Calibri", size=10, bold=True)
        elif ci == 6 and val and "@" in val:  # Email
            c.font = Font(name="Calibri", size=10, color="1565C0", underline="single")
        elif ci == 8 and val:  # Source URL – grey + small
            c.font = Font(name="Calibri", size=9, color="757575", italic=True)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

# Summary sheet
summ = wb.create_sheet("Summary")
summ.column_dimensions["A"].width = 30
summ.column_dimensions["B"].width = 12

summary = [
    ("Dallas-Area School Soccer Coaches", ""),
    ("",""),
    ("Category", "Count"),
    ("High School Coaches", str(sum(1 for r in rows if "HIGH SCHOOL" in r[0]))),
    ("Middle School Coaches", str(sum(1 for r in rows if "MIDDLE SCHOOL" in r[0]))),
    ("Total", str(len(rows))),
    ("",""),
    ("ISDs Covered", ""),
    ("Dallas ISD", ""),
    ("Frisco ISD", ""),
    ("Plano ISD", ""),
    ("Richardson ISD", ""),
    ("Irving ISD", ""),
    ("Cedar Hill ISD", ""),
]

for ri, (a, b) in enumerate(summary, 1):
    ca = summ.cell(row=ri, column=1, value=a)
    cb = summ.cell(row=ri, column=2, value=b)
    if ri == 1:
        ca.font = Font(bold=True, size=13, color=HEADER_BG)
    elif a in ("Category", "ISDs Covered"):
        ca.font = Font(bold=True, size=11, color=HEADER_BG)
        cb.font = Font(bold=True, size=11, color=HEADER_BG)
    elif a == "Total":
        ca.font = Font(bold=True, size=11)
        cb.font = Font(bold=True, size=11)
    else:
        ca.font = Font(size=10)
        cb.font = Font(size=10)

wb.save("/home/node/.openclaw/workspace/hs_ms_soccer_coaches.xlsx")
print("Done!")
